import openpyxl

# Load the Excel file
file_path = '2025-07-14_MDRH_Inventory_Storage_Burn_Rates_V3.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    print("Sheet names:", wb.sheetnames)
    print("\n" + "="*50)
    
    # Examine each sheet
    for sheet_name in wb.sheetnames:
        print(f"\nSHEET: {sheet_name}")
        print("-" * 30)
        
        ws = wb[sheet_name]
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Get the first few rows to understand structure
        print("\nFirst 10 rows of data:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row}: {row_data}")
        
        print("\n" + "="*50)
        
except Exception as e:
    print(f"Error: {e}")
